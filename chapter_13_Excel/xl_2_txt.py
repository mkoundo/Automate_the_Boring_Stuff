#! python3
# xl_2_txt.py
# Author: Michael Koundouros
"""
Write a program that performs the tasks of txt_2_xl.py in reverse order: the program should open a
spreadsheet and write the cells of column A into one text file, the cells of column B into another text file,
and so on.
usage: python xl_2_txt.py file
where file is of xlsx format
"""

import openpyxl
import openpyxl.utils.exceptions
from pathlib import Path
import sys


def export_txt(file):
    # Function to export excel columns to text files.
    try:
        wb = openpyxl.load_workbook(file, data_only=True)                     # Read values instead of formulas
    except FileNotFoundError:
        print(f'{file} not found!')
        sys.exit(1)
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f'{file} file format not supported!')
        sys.exit(1)

    sheet = wb.active
    max_rows = sheet.max_row
    max_cols = sheet.max_column

    for xlcol in range(1, max_cols+1):
        with open(f'{file.stem}{xlcol}.txt', 'a+') as tmpfile:                 # Append to file using a+
            print(f'Writing column {xlcol} to file ...')
            for xlrow in range(1, max_rows):
                tmpfile.write(str(sheet.cell(row=xlrow, column=xlcol).value) + '\n')

    return


def main():
    # Make a list of command line arguments, omitting the [0] element
    # which is the script itself.
    args = sys.argv[1:]

    if len(args) != 1:
        print('usage: python xl_2_txt.py file')
        sys.exit(1)
    else:
        filename = Path(args[0])

    export_txt(filename)
    print('Done!')


if __name__ == '__main__':
    main()
