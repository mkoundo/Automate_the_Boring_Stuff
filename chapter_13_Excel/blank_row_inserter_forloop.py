#! python3
# blank_row_inserter_forloop.py
# Author: Michael Koundouros
"""
Create a program blank_row_inserter.py that takes two integers and a filename string as command line arguments.
Let’s call the first integer N and the second integer M. Starting at row N, the program should insert M blank rows
into the spreadsheet.

You can write this program by reading in the contents of the spreadsheet. Then, when writing out the new spreadsheet,
use a for loop to copy the first N lines. For the remaining lines, add M to the row number in the output spreadsheet.
python blank_row_inserter_forloop.py row blank_rows file
"""


import openpyxl
import openpyxl.utils.exceptions
from pathlib import Path
import sys


def row_insert(n_row, m_blank_rows, file):
    # Function to insert blank rows in excel spreadsheet
    try:
        wb = openpyxl.load_workbook(file, data_only=True)                     # Read values instead of formulas
    except FileNotFoundError:
        print(f'{file} not found!')
        sys.exit(1)
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f'{file} file format not supported!')
        sys.exit(1)

    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    print(f'Inserting {m_blank_rows} rows at row {n_row}...')
    sheet = wb.active
    max_rows = sheet.max_row
    max_cols = sheet.max_column

    ranges = ((1, n_row, 0),                                                  # Rows 1 to first blank row
              (n_row + m_blank_rows, max_rows + 1 + m_blank_rows, 1))         # Rows last blank row to last row

    for range_value in ranges:
        for row in range(range_value[0], range_value[1]):
            for col in range(1, max_cols+1):
                new_sheet.cell(row=row, column=col).value = \
                    sheet.cell(row=row-(range_value[2]*m_blank_rows), column=col).value

    wb.close()
    new_wb.save(f'{file.stem}_amended.xlsx')

    return


def main():
    # Make a list of command line arguments, omitting the [0] element
    # which is the script itself.
    args = sys.argv[1:]

    if len(args) != 3:
        print('usage: python blank_row_inserter_forloop.py row blank_rows file')
        sys.exit(1)
    else:
        filename = Path(args[2])
        try:
            n_integer = int(args[0])
            m_integer = int(args[1])
            openpyxl.load_workbook(filename)
        except ValueError:
            print('Invalid arguments given!')
            sys.exit(1)

    row_insert(n_integer, m_integer, filename)
    print('Done!')


if __name__ == '__main__':
    main()
