#! python3
# multiplication_table.py
# Author: Michael Koundouros
"""
Create a program multiplicationTable.py that takes a number N from the command line and creates an N×N
multiplication table in an Excel spreadsheet. For example, when the program is run like this:
python multiplication_table.py 6
"""


import openpyxl
from openpyxl.styles import Font
import sys


def excel_multiplication_table(mult_int):
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Create row and column labels
    for row in range(1, mult_int+1):
        sheet.cell(row+1, 1).value = row
        sheet.cell(row + 1, 1).font = Font(bold=True)
    for column in range( 1, mult_int+1):
        sheet.cell(1, column+1).value = column
        sheet.cell(1, column + 1).font = Font(bold=True)

    # Create multiplication table
    for row in range(1, mult_int+1):
        for column in range(1, mult_int+1):
            sheet.cell(row+1, column+1).value = row * column

    wb.save(f'{mult_int}x{mult_int}.xlsx')

    return


def main():
    # Make a list of command line arguments, omitting the [0] element
    # which is the script itself.
    args = sys.argv[1:]

    if len(args) != 1:
        print('usage: python multiplication_table.py integer')
        sys.exit(1)
    else:
        multiplication_integer = int(args[0])

    excel_multiplication_table(multiplication_integer)
    print('Done!')


if __name__ == '__main__':
    main()
