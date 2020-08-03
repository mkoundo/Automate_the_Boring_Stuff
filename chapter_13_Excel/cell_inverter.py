#! python3
# cell_inverter.py
# Author: Michael Koundouros
"""
Write a program to invert the row and column of the cells in the spreadsheet. For example, the value at row 5,
column 3 will be at row 3, column 5 (and vice versa). This should be done for all cells in the spreadsheet.
"""


import openpyxl
import openpyxl.utils.exceptions
from pathlib import Path
import sys


def invert(file):
    # Transpose excel spreadsheet cells
    try:
        wb = openpyxl.load_workbook(file, data_only=True)                     # Read values instead of formulas
    except FileNotFoundError:
        print(f'{file} not found!')
        sys.exit(1)
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f'{file} file format not supported!')
        sys.exit(1)

    sheet = wb.active
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    max_rows = sheet.max_row
    max_cols = sheet.max_column

    if max_rows > 16384:                                                      # Excel has a 16384 column limit
        print('There are more rows than available columns. Exiting...')
        return

    print(f'Transposing cells...')

    for row in range(1, max_rows + 1):
        for col in range(1, max_cols + 1):
            new_sheet.cell(row=col, column=row).value = sheet.cell(row=row, column=col).value

    wb.close()
    new_wb.save(f'{file.stem}_amended.xlsx')

    return


def main():
    # Make a list of command line arguments, omitting the [0] element
    # which is the script itself.
    args = sys.argv[1:]

    if len(args) != 1:
        print('usage: python cell_inverter.py file')
        sys.exit(1)
    else:
        filename = Path(args[0])

    invert(filename)
    print('Done!')


if __name__ == '__main__':
    main()
