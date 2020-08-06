#! python3
# xl2csv.py
# Author: Michael Koundouros
"""
Excel can save a spreadsheet to a CSV file with a few mouse clicks, but if you had to convert hundreds of Excel
files to CSVs, it would take hours of clicking. Using the openpyxl module from Chapter 12, write a program that reads
all the Excel files in the current working directory and outputs them as CSV files.

A single Excel file might contain multiple sheets; you’ll have to create one CSV file per sheet. The filenames of the
CSV files should be <excel filename>_<sheet title>.csv, where <excel filename> is the filename of the Excel file
without the file extension (for example, 'spam_data', not 'spam_data.xlsx') and <sheet title> is the string from the
Worksheet object’s title variable.
"""


import openpyxl
import csv
from pathlib import Path

# Convert all spreadsheets in current working directory to csv files
xl_list = list(Path.cwd().glob('*.xlsx'))
print('Writing files...')

for xl_file in xl_list:
    wb = openpyxl.load_workbook(xl_file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        output_csv = open(f'{xl_file.stem}_{sheet}.csv', 'w', newline='')
        print(f'{xl_file.stem}_{sheet}.csv')

        for row_num in range(1, ws.max_row + 1):
            row_data = []
            for cell in range(0, ws.max_column):
                cell_data = ws[row_num][cell].value
                row_data.append(cell_data)
            output_writer = csv.writer(output_csv)
            output_writer.writerow(row_data)
        output_csv.close()

print('Done!')
